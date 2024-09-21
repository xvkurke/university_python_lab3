import csv
from openpyxl import Workbook
from datetime import datetime

def calculate_age(birthdate):
    today = datetime.today()
    return today.year - birthdate.year - ((today.month, today.day) < (birthdate.month, birthdate.day))

def write_to_excel(data):
    workbook = Workbook()
    sheets = {
        'all': workbook.active,
        'younger_18': workbook.create_sheet('younger_18'),
        '18-45': workbook.create_sheet('18-45'),
        '45-70': workbook.create_sheet('45-70'),
        'older_70': workbook.create_sheet('older_70')
    }
    sheets['all'].title = 'all'
    
    for sheet in sheets.values():
        sheet.append(['№', 'Прізвище', 'Ім’я', 'По батькові', 'Дата народження', 'Вік'])
    
    for idx, row in enumerate(data):
        try:
            birthdate = datetime.strptime(row[4], '%Y-%m-%d')
            age = calculate_age(birthdate)
            new_row = [idx+1, *row[:4], row[4], age]
            sheets['all'].append(new_row)
            if age < 18:
                sheets['younger_18'].append(new_row)
            elif 18 <= age <= 45:
                sheets['18-45'].append(new_row)
            elif 45 < age <= 70:
                sheets['45-70'].append(new_row)
            else:
                sheets['older_70'].append(new_row)
        except Exception as e:
            print(f"Error processing row {idx+1}: {e}")

    try:
        workbook.save('employees.xlsx')
        print("Ok")
    except Exception as e:
        print(f"Unable to create XLSX file: {e}")

# Read CSV and process
try:
    with open('employees.csv', newline='', encoding='utf-8') as file:
        reader = csv.reader(file)
        next(reader)  # Skip the header
        data = list(reader)
        write_to_excel(data)
except FileNotFoundError:
    print("CSV file not found.")
